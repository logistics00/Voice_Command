"""
make_calling.py
---------------
Scans Voice_Command.ahk and all lib/Project/*.ahk files.
Builds Calling.xlsx with 5 columns:
  Col 1  CallerScript    - script that contains the call
  Col 2  FunctionName    - name of the called function
  Col 3  CallerLineNr    - line number of the call
  Col 4  DefinedInScript - script where the function is defined
  Col 5  DefinedAtLineNr - line number of the function definition

Only user-defined functions are recorded (built-ins are excluded automatically
because they never appear in the function-definition map).
"""

import re
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = r"D:\Sync\AHK\Connie_Marks\Voice_Command\Voice_Command_NMS"

FILES = {
    "Voice_Command.ahk":       os.path.join(ROOT, "Voice_Command.ahk"),
    "Voice_Command_UI.ahk":    os.path.join(ROOT, "lib", "Project", "Voice_Command_UI.ahk"),
    "Voice_Command_Bridge.ahk":os.path.join(ROOT, "lib", "Project", "Voice_Command_Bridge.ahk"),
    "Voice_Command_Core.ahk":  os.path.join(ROOT, "lib", "Project", "Voice_Command_Core.ahk"),
    "Voice_Command_Utils.ahk": os.path.join(ROOT, "lib", "Project", "Voice_Command_Utils.ahk"),
}

# AHK keywords that look like function definitions but are not
AHK_KEYWORDS = {
    'if', 'else', 'while', 'loop', 'for', 'switch', 'case',
    'try', 'catch', 'return', 'break', 'continue',
    'global', 'local', 'static', 'throw',
}

# Matches a function/method definition line (after stripping leading whitespace).
# Requires '{' on the same line to distinguish definitions from standalone calls:
#   FunctionName(params) {
DEF_PATTERN = re.compile(
    r'^([A-Za-z_][A-Za-z0-9_]*)\s*\([^)]*\)\s*\{'
)

# Matches any function call: word immediately followed by (
CALL_PATTERN = re.compile(r'\b([A-Za-z_][A-Za-z0-9_]*)\s*\(')


def read_lines(path):
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        return f.readlines()


def strip_inline_comment(line):
    """Remove AHK inline comment (;) that is not inside a string."""
    in_str = False
    str_char = None
    for i, ch in enumerate(line):
        if not in_str:
            if ch in ('"', "'"):
                in_str = True
                str_char = ch
            elif ch == ';':
                return line[:i]
        else:
            if ch == str_char:
                in_str = False
    return line


# ── Pass 1: collect all function definitions ──────────────────────────────────
# func_defs[name_lower] = (script_name, line_number, original_case_name)
func_defs: dict[str, tuple[str, int, str]] = {}

for script_name, file_path in FILES.items():
    lines = read_lines(file_path)
    for line_nr, raw in enumerate(lines, 1):
        stripped = raw.strip()
        if stripped.startswith(';'):          # full-line comment
            continue
        m = DEF_PATTERN.match(stripped)
        if m:
            name = m.group(1)
            if name.lower() not in AHK_KEYWORDS:
                key = name.lower()
                if key not in func_defs:      # first definition wins
                    func_defs[key] = (script_name, line_nr, name)

print(f"Found {len(func_defs)} user-defined functions:")
for k, v in sorted(func_defs.items()):
    print(f"  {v[2]:40s}  {v[0]}:{v[1]}")


# ── Pass 2: find all calls to user-defined functions ─────────────────────────
rows = []   # (caller_script, func_name, caller_line, def_script, def_line)

for script_name, file_path in FILES.items():
    lines = read_lines(file_path)
    for line_nr, raw in enumerate(lines, 1):
        stripped = raw.strip()
        if stripped.startswith(';'):
            continue

        # Remove inline comment before scanning
        code = strip_inline_comment(stripped)

        # Determine if this line itself IS a function definition
        def_match = DEF_PATTERN.match(stripped)
        is_def_of = def_match.group(1).lower() if def_match else None

        for call_m in CALL_PATTERN.finditer(code):
            called = call_m.group(1)
            called_lower = called.lower()

            if called_lower not in func_defs:
                continue                       # not a user-defined function

            # Skip the definition line itself (avoids recording "FunctionName(" in its own header)
            if is_def_of == called_lower:
                continue

            def_script, def_line, def_original = func_defs[called_lower]
            rows.append((
                script_name,    # Col 1
                def_original,   # Col 2  (use canonical casing from definition)
                line_nr,        # Col 3
                def_script,     # Col 4
                def_line,       # Col 5
            ))

# Sort by caller script (file order), then line number
file_order = list(FILES.keys())
rows.sort(key=lambda r: (file_order.index(r[0]), r[2]))

print(f"\nFound {len(rows)} function calls.\n")


# ── Pass 3: write Excel ───────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Function Calls"

HEADERS = ["CallerScript", "FunctionName", "CallerLineNr", "DefinedInScript", "DefinedAtLineNr"]
HDR_FILL = PatternFill("solid", fgColor="4472C4")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HDR_ALIGN = Alignment(horizontal="center", vertical="center")

# Alternating row colors
FILL_ODD  = PatternFill("solid", fgColor="DCE6F1")
FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")

ws.row_dimensions[1].height = 20

for col, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = HDR_ALIGN

for row_idx, row_data in enumerate(rows, 2):
    fill = FILL_ODD if row_idx % 2 == 0 else FILL_EVEN
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")

# Column widths
COL_WIDTHS = [30, 38, 14, 30, 16]
for col, width in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:E{len(rows) + 1}"

output_path = os.path.join(ROOT, "Calling.xlsx")
wb.save(output_path)
print(f"Saved: {output_path}  ({len(rows)} data rows)")
