import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = [
    # Voice_Command.ahk  — entry point only, no function definitions
    # ── Voice_Command_UI.ahk ──────────────────────────────────────────
    ("Voice_Command_UI.ahk", "SetupTrayMenu", "()",
     "Build system tray menu with all menu items and set default action."),
    ("Voice_Command_UI.ahk", "ShowCommandManagerMenu", "(*)",
     "Tray menu wrapper: calls ShowCommandManagerGui()."),
    ("Voice_Command_UI.ahk", "ShowMicSettingsMenu", "(*)",
     "Tray menu wrapper: calls ShowMicrophoneSettingsGui()."),
    ("Voice_Command_UI.ahk", "ReloadCommandsMenu", "(*)",
     "Tray menu: reload commands from INI and rebuild SAPI grammar."),
    ("Voice_Command_UI.ahk", "EditIniMenu", "(*)",
     "Tray menu: open INI file in Notepad."),
    ("Voice_Command_UI.ahk", "ToggleListeningMenu", "(*)",
     "Tray menu wrapper: calls ToggleListening()."),
    ("Voice_Command_UI.ahk", "ToggleLoggingMenu", "()",
     "Tray menu: toggle blnLogEnabled on/off and persist to INI."),
    ("Voice_Command_UI.ahk", "ExitMenu", "(*)",
     "Tray menu: call CleanupVoice() for clean exit (guarded against re-entry)."),
    ("Voice_Command_UI.ahk", "UpdateTrayIcon", "()",
     "Set tray icon and tooltip based on blnListening / blnCommandsEnabled; calls UpdateStatusCircle."),
    ("Voice_Command_UI.ahk", "ToggleListening", "()",
     "Toggle SAPI grammars on/off (F1); update tray icon and tooltip."),
    ("Voice_Command_UI.ahk", "GetStatusLabel", "()",
     "Return language label (EN/NL) for Vosk/Whisper modes; empty string for SAPI."),
    ("Voice_Command_UI.ahk", "GetStatusColor", "()",
     "Return hex color for current voice mode and listening state."),
    ("Voice_Command_UI.ahk", "CreateStatusCircle", "()",
     "Create colored circle overlay in upper-right corner of screen."),
    ("Voice_Command_UI.ahk", "UpdateStatusCircle", "()",
     "Destroy and recreate status circle with current color and label."),
    ("Voice_Command_UI.ahk", "ShowCircle", "(circleSize, circleColor, xPos, yPos, strLabel)",
     "Create elliptical GUI window with optional center label; return GUI object."),
    ("Voice_Command_UI.ahk", "ShowCommandManagerGui", "()",
     "Build and show Command Manager GUI: hotkey editor, input fields, ListView."),
    ("Voice_Command_UI.ahk", "UpdateHotkey", "(cb, edt, token, type)",
     "Checkbox click: add/remove modifier token, validate, save hotkey to INI and re-register."),
    ("Voice_Command_UI.ahk", "UpdateHotkeyEdtField", "(edt, type)",
     "Edit field change: validate new hotkey string, save to INI and re-register."),
    ("Voice_Command_UI.ahk", "CheckHotkey", "(hotkey, message)",
     "Test-register a hotkey to validate it; show MsgBox on error; return true/false."),
    ("Voice_Command_UI.ahk", "RefreshCommandList", "()",
     "Clear and repopulate ListView with all entries from mapCommands."),
    ("Voice_Command_UI.ahk", "CommandListClick", "(ctrl, info)",
     "ListView single-click: populate input fields with selected row data."),
    ("Voice_Command_UI.ahk", "CommandListDoubleClick", "(ctrl, info)",
     "ListView double-click: delegates to CommandListClick."),
    ("Voice_Command_UI.ahk", "AddCommand", "(*)",
     "Read input fields, add command to mapCommands + INI, refresh list, rebuild grammar."),
    ("Voice_Command_UI.ahk", "EditCommand", "(*)",
     "Update selected command in mapCommands + INI (rename if phrase changed), rebuild grammar."),
    ("Voice_Command_UI.ahk", "DeleteCommand", "(*)",
     "Confirm and delete selected command from mapCommands + INI, rebuild grammar."),
    ("Voice_Command_UI.ahk", "ClearFields", "(*)",
     "Clear all Command Manager input fields and reset intSelectedRow to 0."),
    ("Voice_Command_UI.ahk", "CmdManagerClose", "(*)",
     "Destroy Command Manager GUI on close event; reset goo to empty."),
    ("Voice_Command_UI.ahk", "ShowMicrophoneSettingsGui", "(blnForceSelection := false)",
     "Build and show Microphone Settings GUI with list, level meter and confidence slider."),
    ("Voice_Command_UI.ahk", "ThresholdSliderChange", "(ctrl, *)",
     "Update threshold percentage label when confidence slider moves."),
    ("Voice_Command_UI.ahk", "MicListClick", "(ctrl, info)",
     "ListView click: switch active microphone immediately for real-time testing."),
    ("Voice_Command_UI.ahk", "SaveMicSettings", "(*)",
     "Save selected mic index/name and confidence settings to INI; close GUI."),
    ("Voice_Command_UI.ahk", "MicSettingsClose", "(*)",
     "Stop audio monitor, destroy Microphone Settings GUI."),
    ("Voice_Command_UI.ahk", "TestMicRecognition", "(*)",
     "Start 5-second mic test; set blnMicTestMode and schedule ResetMicTest."),
    ("Voice_Command_UI.ahk", "ResetMicTest", "()",
     "Timer callback: reset blnMicTestMode and update test result label after 5 s."),
    ("Voice_Command_UI.ahk", "StartAudioLevelMonitor", "()",
     "Start 100 ms repeating timer that calls UpdateAudioLevel."),
    ("Voice_Command_UI.ahk", "StopAudioLevelMonitor", "()",
     "Stop the UpdateAudioLevel timer (interval 0)."),
    ("Voice_Command_UI.ahk", "UpdateAudioLevel", "()",
     "Read SAPI AudioLevel (0-100) and update progress bar and label."),

    # ── Voice_Command_Utils.ahk ──────────────────────────────────────
    ("Voice_Command_Utils.ahk", "SetupBuiltInCommands", "()",
     "Populate mapControlCommands (start/pause) and mapBuiltInCommands (list/stop/start)."),
    ("Voice_Command_Utils.ahk", "CreateDefaultIni", "()",
     "Write default [Settings] and sample [Commands] to INI when file is missing."),
    ("Voice_Command_Utils.ahk", "LoadCommandsFromIni", "()",
     "Read [Commands] section from INI; return Map of lowercase phrase -> action string."),
    ("Voice_Command_Utils.ahk", "LoadConfidenceSettings", "()",
     "Read confidenceThreshold and showConfidence from INI into globals."),
    ("Voice_Command_Utils.ahk", "FFL", "(strFunc, intLine)",
     "Format 'FuncName(line)' padded to 60 chars for consistent log prefixes."),
    ("Voice_Command_Utils.ahk", "LogMsg", "(strMessage, display := 0)",
     "Append timestamped message to log file if intLoggingType matches display level."),
    ("Voice_Command_Utils.ahk", "CleanupVoice", "(exitReason, exitCode)",
     "Disconnect bridge, stop audio monitor, destroy circle overlay, exit app."),
    ("Voice_Command_Utils.ahk", "ExampleFunction", "()",
     "Demo function: show MsgBox; usable as voice command target via type=Function."),
    ("Voice_Command_Utils.ahk", "StartCommands", "()",
     "Enable main SAPI grammar rule; show tooltip; update tray icon."),
    ("Voice_Command_Utils.ahk", "PauseCommands", "()",
     "Disable main SAPI grammar rule; show tooltip; update tray icon."),

    # ── Voice_Command_Core.ahk ───────────────────────────────────────
    ("Voice_Command_Core.ahk", "InitializeVoiceRecognition", "()",
     "Init SAPI recognizer, verify mic, build dual grammars, start recognition, show ready MsgBox."),
    ("Voice_Command_Core.ahk", "VerifyMicrophone", "()",
     "Check saved mic index/name against available inputs; return 'OK' or 'SHOW_GUI'."),
    ("Voice_Command_Core.ahk", "BuildGrammarFile", "(mapUserCmds, mapBuiltIn, strFilePath)",
     "Write SAPI XML grammar file from built-in + user command phrases."),
    ("Voice_Command_Core.ahk", "BuildControlGrammarFile", "(strFilePath)",
     "Write SAPI XML grammar file for control commands only (start/pause)."),
    ("Voice_Command_Core.ahk", "RebuildGrammar", "()",
     "Disable main grammar, rebuild XML from current mapCommands, reload and re-enable."),
    ("Voice_Command_Core.ahk", "VoiceEventSink.__Call", "(strMethod, args)",
     "Route SAPI COM events to Recognition / Hypothesis / FalseRecognition handlers."),
    ("Voice_Command_Core.ahk", "VoiceEventSink.HandleRecognition", "(args)",
     "Filter recognized phrase by confidence threshold; dispatch to ControlHandler or VoiceHandler."),
    ("Voice_Command_Core.ahk", "VoiceEventSink.LogHypothesis", "(args)",
     "Log hypothesis text; show tooltip when sapiSpeakMode=1."),
    ("Voice_Command_Core.ahk", "VoiceEventSink.LogFalseRecognition", "(args)",
     "Log false recognition text; show tooltip when sapiSpeakMode=1."),
    ("Voice_Command_Core.ahk", "ControlHandler", "(strCommand)",
     "Handle start/pause control commands: enable/disable main grammar and update tray."),
    ("Voice_Command_Core.ahk", "VoiceHandler", "(result)",
     "Look up recognized text in built-in and user command maps; call ExecuteAction."),
    ("Voice_Command_Core.ahk", "ExecuteAction", "(strActionData)",
     "Parse group|type|target string and execute run/send/winclose/mouse/builtin/function action."),
    ("Voice_Command_Core.ahk", "ExecuteMouseAction", "(strTarget)",
     "Parse and execute click or double-click on a named window."),
    ("Voice_Command_Core.ahk", "ShowCommandList", "()",
     "Display MsgBox listing all control, built-in and user commands with types."),

    # ── Voice_Command_Bridge.ahk ─────────────────────────────────────
    ("Voice_Command_Bridge.ahk", "BridgeKillOrphan", "()",
     "Connect to port 7891; if open send QUIT so old bridge shuts down cleanly."),
    ("Voice_Command_Bridge.ahk", "BridgeInit", "()",
     "Read localLanguage, kill orphan, launch bridge.py, retry TCP connect up to 20x."),
    ("Voice_Command_Bridge.ahk", "BridgeConnect", "()",
     "Open Winsock TCP socket to bridge, set non-blocking, start 50 ms receive timer."),
    ("Voice_Command_Bridge.ahk", "BridgeSend", "(strMsg)",
     "Encode msg as UTF-8 and send over TCP socket (newline appended automatically)."),
    ("Voice_Command_Bridge.ahk", "BridgeReceiveLoop", "()",
     "50 ms timer: recv data into buffer, split on newline, dispatch complete lines."),
    ("Voice_Command_Bridge.ahk", "BridgeHandleMessage", "(strMsg)",
     "Dispatch TEXT: (type text), STATUS: (log), ERROR: (tooltip + log) from bridge."),
    ("Voice_Command_Bridge.ahk", "BridgeDisconnect", "()",
     "Send QUIT, wait for bridge process, close socket, cleanup Winsock, kill PID."),
    ("Voice_Command_Bridge.ahk", "CycleVoiceMode", "(*)",
     "F3: cycle SAPI -> Vosk -> Whisper -> SAPI; blocked when listening is OFF."),
    ("Voice_Command_Bridge.ahk", "SwitchToVosk", "()",
     "Pause both SAPI grammars; send MODE:vosk to bridge; update circle."),
    ("Voice_Command_Bridge.ahk", "SwitchToWhisper", "()",
     "Pause both SAPI grammars; send MODE:whisper to bridge; update circle."),
    ("Voice_Command_Bridge.ahk", "SwitchToSapi", "()",
     "Send MODE:sapi; restore SAPI grammars; update circle."),
    ("Voice_Command_Bridge.ahk", "ToggleLanguage", "(*)",
     "F4: toggle speakLanguage default/special; send LANG: to bridge; update circle."),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Function Overview"

# ── Styles ───────────────────────────────────────────────────────────
hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
hdr_fill  = PatternFill("solid", fgColor="1F4E79")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

grp_fills = {
    "Voice_Command_UI.ahk":     PatternFill("solid", fgColor="DEEAF1"),
    "Voice_Command_Utils.ahk":  PatternFill("solid", fgColor="E2EFDA"),
    "Voice_Command_Core.ahk":   PatternFill("solid", fgColor="FFF2CC"),
    "Voice_Command_Bridge.ahk": PatternFill("solid", fgColor="FCE4D6"),
}

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

data_align = Alignment(vertical="top", wrap_text=True)

# ── Header ───────────────────────────────────────────────────────────
headers = ["Script", "Function", "Parameters", "Description"]
ws.append(headers)
for col, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.font  = hdr_font
    cell.fill  = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

ws.row_dimensions[1].height = 22

# ── Data rows ────────────────────────────────────────────────────────
for r_idx, (script, func, params, desc) in enumerate(rows, start=2):
    ws.cell(r_idx, 1, script)
    ws.cell(r_idx, 2, func)
    ws.cell(r_idx, 3, params)
    ws.cell(r_idx, 4, desc)

    fill = grp_fills.get(script, PatternFill())
    for c in range(1, 5):
        cell = ws.cell(r_idx, c)
        cell.fill      = fill
        cell.border    = border
        cell.alignment = data_align
        cell.font      = Font(name="Calibri", size=10)

    ws.row_dimensions[r_idx].height = 30

# ── Column widths ────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 70

# ── Freeze header row ────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Auto-filter ──────────────────────────────────────────────────────
ws.auto_filter.ref = f"A1:D{len(rows)+1}"

out = r"D:\Sync\AHK\Connie_Marks\Voice_Command\Voice_Command_NMS\Function_Overview.xlsx"
wb.save(out)
print(f"Saved: {out}  ({len(rows)} functions)")
